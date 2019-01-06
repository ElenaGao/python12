# -*- coding:utf-8 _*-
""" 
@author:mongo
@file: do_excel.py 
@version:
@time: 2018/12/17 
@email:3126972006@qq.com
@function： 
"""

import json

import openpyxl

from common.request import Request
from common.config import ConfigLoader
from common import contants


class Case:
    """
    测试用例封装类
    """

    def __init__(self):
        self.case_id = None
        self.url = None
        self.data = None
        self.title = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None


class DoExcel:
    """
    excel 操作封装类
    """

    def __init__(self, file_name):  # DoExcel初始化函数
        try:
            self.file_name = file_name
            # 打开一个excel文件，返回一个workbook对象实例，把它定义为DoExcel的属性，以便于在这个类的其他地方使用
            self.workbook = openpyxl.load_workbook(filename=file_name)
        except FileNotFoundError as e:
            # 文件未找到异常处理
            print('{0} not found, please check file path'.format(file_name))
            raise e

    def get_cases(self, sheet_name):  # 根据sheet名称，获取在这个sheet里面的所有测试用例数据
        sheet = self.workbook[sheet_name]  # 根据sheet名称获取sheet对象实例
        max_row = sheet.max_row  # 获取sheet最大行数
        cases = []  # 定义一个列表，用来存放即将要放进去的测试用例
        for r in range(2, max_row + 1):  # for 循环，从第二行开始遍历
            case = Case()  # 实例化一个case对象，用来存放测试数据
            case.case_id = sheet.cell(row=r, column=1).value  # 取第r行，第1格的值
            case.title = sheet.cell(row=r, column=2).value  # 取第r行，第4格的值
            case.url = sheet.cell(row=r, column=3).value  # 取第r行，第2格的值
            case.data = sheet.cell(row=r, column=4).value  # 取第r行，第3格的值
            case.method = sheet.cell(row=r, column=5).value  # 取第r行，第5格的值
            case.expected = sheet.cell(row=r, column=6).value  # 取第r行，第6格的值
            cases.append(case)  # 将case放到cases 列表里面

        return cases  # for 循环结束后返回cases列表

    def get_sheet_names(self):
        return self.workbook.sheetnames  # Returns the list of the names of worksheets in this workbook.

    # 根据sheet_name定位到sheet,再根据case_id定位到当前行，再定位到actual的单元格，进行写入并保存
    def write_result_by_case_id(self,sheet_name,case_id,actual,result):
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        for r in range(2,max_row+1):
            case_id_r = sheet.cell(r,1).value
            if case_id_r == case_id:
                sheet.cell(r,7).value = actual
                sheet.cell(r,8).value = result
                self.workbook.save(filename=self.file_name)
                break

if __name__ == '__main__':
    print('comming')
    #  测试一下DoExcel类
    do_excel = DoExcel(contants.case_file)  # 实例化一个DoExcel对象
    sheet_names = do_excel.get_sheet_names()  # 获取Excel文件中全部sheet名称，返回为一个列表
    print("sheet名称列表：",sheet_names)
    cases_list = ['login', 'register'] # 定义一个执行测试用例的列表

    #  根据sheet_names 分别取到全部用例来执行
    for sheet_name in sheet_names:
        # 获取login的所有测试用例，返回的是一个cases列表，列表里面是多个Case类实例，每个实例的属性对应Excel表中每列的值
        if sheet_name in cases_list: #如果当前的这个sheet_name不在可执行的case_list里面，就不执行
            cases = do_excel.get_cases(sheet_name)
            print(sheet_name+' 测试用例个数：', len(cases))
            for case in cases:  # 遍历测试用例列表，每进for一次，就取一个case实例
                print("case信息：", case.__dict__)  # 打印case信息  对象的内置的私有方法
                # data = eval(case.data)  # Excel里面取到data是一个字符串，使用eval函数将字符串转换成字典
                data = json.loads(case.data)
                config = ConfigLoader()
                url_pre = config.get('api','url_pre')
                url = url_pre + case.url
                resp = Request(method=case.method, url=url, data=data)  # 通过封装的Request类来完成接口的调用
                print('status_code:', resp.get_status_code())  # 打印响应码
                resp_dict = resp.get_json()  # 获取请求响应，字典
                # 通过json.dumps函数将字典转换成格式化后的字符串
                resp_text = json.dumps(resp_dict, ensure_ascii=False, indent=4)
                print('response: ', resp_text)  # 打印响应

                if case.expected == resp.get_text():
                    print("result : PASS")
                    do_excel.write_result_by_case_id(sheet_name=sheet_name,case_id=case.case_id,
                                                     actual=resp.get_text(),result='PASS')
                else:
                    print("result : FAIL")
                    do_excel.write_result_by_case_id(sheet_name=sheet_name,case_id=case.case_id,
                                                     actual=resp.get_text(),result='FAIL')